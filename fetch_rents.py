#!/usr/bin/env python3
"""
fetch_rents.py
Downloads HUD Small Area Fair Market Rents (SAFMR) for LA/OC area zip codes
and outputs rents.json keyed by zip code with 3BR and 4BR FMR values.

Usage:
    python3 fetch_rents.py

Output:
    rents.json — { "90210": {"fmr3br": 4200, "fmr4br": 5100}, ... }
"""
import json, os, re, statistics, sys
import requests

os.chdir(os.path.dirname(os.path.abspath(__file__)))
from market_config import get_market, market_file

market = get_market()
listings_file = market_file("listings.js", market)
output_file = market_file("rents.json", market)

# ── Step 1: Extract unique zip codes from listings.js ──
print(f"\n📋 Step 1: Loading zip codes from {listings_file}...")
if not os.path.exists(listings_file):
    print(f"   ❌ {listings_file} not found — run: python3 listings_build.py")
    sys.exit(1)

with open(listings_file, "r") as f:
    raw = f.read()
start = raw.index("[")
end = raw.rindex("]") + 1
listings = json.loads(raw[start:end])

our_zips = set()
for l in listings:
    z = str(l.get("zip", "")).strip()
    if z and len(z) == 5 and z.isdigit():
        our_zips.add(z)

print(f"   Found {len(our_zips)} unique zip codes from {len(listings):,} listings")

# ── Step 2: Download HUD SAFMR XLSX ──
print("\n📥 Step 2: Downloading HUD FY2025 SAFMR data...")

SAFMR_URLS = [
    "https://www.huduser.gov/portal/datasets/fmr/fmr2025/fy2025_safmrs.xlsx",
    "https://www.huduser.gov/portal/datasets/fmr/fmr2025/FY25_FMRs_revised.xlsx",
]

CACHE_FILE = "safmr_cache.xlsx"
xlsx_path = None

# Use cache if recent
if os.path.exists(CACHE_FILE):
    age_hours = (os.time() if hasattr(os, 'time') else __import__('time').time() - os.path.getmtime(CACHE_FILE)) / 3600
    if age_hours < 168:  # 7 days
        print(f"   Using cached {CACHE_FILE} ({age_hours:.0f}h old)")
        xlsx_path = CACHE_FILE

if not xlsx_path:
    for url in SAFMR_URLS:
        try:
            print(f"   Trying: {url}")
            resp = requests.get(url, timeout=60, headers={
                "User-Agent": "Mozilla/5.0 (Python/fetch_rents.py)"
            })
            if resp.status_code == 200 and len(resp.content) > 10000:
                with open(CACHE_FILE, "wb") as f:
                    f.write(resp.content)
                print(f"   ✅ Downloaded {len(resp.content):,} bytes")
                xlsx_path = CACHE_FILE
                break
            else:
                print(f"   ⚠️  HTTP {resp.status_code} (size: {len(resp.content)})")
        except Exception as e:
            print(f"   ⚠️  Failed: {e}")

if not xlsx_path:
    print("   ❌ Could not download SAFMR data from any URL.")
    print("   Please manually download from: https://www.huduser.gov/portal/datasets/fmr/smallarea/index.html")
    print("   Save as safmr_cache.xlsx in this directory and re-run.")
    sys.exit(1)

# ── Step 3: Parse XLSX for 3BR and 4BR FMR by zip code ──
print("\n📊 Step 3: Parsing SAFMR data...")
try:
    from openpyxl import load_workbook
except ImportError:
    print("   ❌ openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

wb = load_workbook(xlsx_path, read_only=True, data_only=True)
ws = wb.active

# Find header row — look for columns containing zip, fmr, 3br, 4br etc.
headers = []
header_row = None
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
    cells = [str(c.value or "").strip().lower() for c in row]
    # Look for a row that has "zip" and "fmr" or "safmr" in column names
    if any("zip" in c for c in cells) and any("fmr" in c or "rent" in c for c in cells):
        headers = cells
        header_row = row_idx
        break

if not header_row:
    # Try first row as header
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    headers = [str(c.value or "").strip().lower() for c in first_row]
    header_row = 1

print(f"   Header row: {header_row}")
print(f"   Columns: {headers[:15]}...")

# Find zip code column and FMR columns
zip_col = None
fmr3_col = None
fmr4_col = None

for i, h in enumerate(headers):
    hl = h.lower().replace("\n", " ")
    if "zip" in hl and zip_col is None:
        zip_col = i
    # Match base SAFMR columns (e.g. "safmr 3br") but NOT payment standard variants
    if "payment" in hl or "90%" in hl or "110%" in hl:
        continue
    if re.search(r'(safmr|fmr).*3\s*b', hl) or re.search(r'3\s*b.*rent', hl):
        if fmr3_col is None:
            fmr3_col = i
    if re.search(r'(safmr|fmr).*4\s*b', hl) or re.search(r'4\s*b.*rent', hl):
        if fmr4_col is None:
            fmr4_col = i

# Fallback: look for columns by position if specific names not found
if fmr3_col is None or fmr4_col is None:
    for i, h in enumerate(headers):
        hl = h.lower().replace("\n", " ")
        if "payment" in hl or "90%" in hl or "110%" in hl:
            continue
        if fmr3_col is None and "3" in hl and ("br" in hl or "bed" in hl):
            fmr3_col = i
        if fmr4_col is None and "4" in hl and ("br" in hl or "bed" in hl):
            fmr4_col = i

if zip_col is None:
    print(f"   ❌ Could not find ZIP column in headers: {headers}")
    sys.exit(1)

print(f"   ZIP col: {zip_col} ({headers[zip_col]})")
if fmr3_col is not None:
    print(f"   3BR FMR col: {fmr3_col} ({headers[fmr3_col]})")
if fmr4_col is not None:
    print(f"   4BR FMR col: {fmr4_col} ({headers[fmr4_col]})")

if fmr3_col is None:
    print(f"   ❌ Could not find 3BR FMR column. Headers: {headers}")
    sys.exit(1)

# Read data rows
all_rents = {}
total_rows = 0
for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
    total_rows += 1
    try:
        raw_zip = str(row[zip_col] or "").strip()
        # Extract 5-digit zip from various formats (e.g., "ZCTA5 90210", "90210.0")
        zip_match = re.search(r'(\d{5})', raw_zip)
        if not zip_match:
            continue
        zipcode = zip_match.group(1)

        fmr3 = row[fmr3_col] if fmr3_col is not None else None
        fmr4 = row[fmr4_col] if fmr4_col is not None else None

        if fmr3 is not None:
            try:
                fmr3 = int(float(str(fmr3).replace(",", "").replace("$", "")))
            except (ValueError, TypeError):
                fmr3 = None

        if fmr4 is not None:
            try:
                fmr4 = int(float(str(fmr4).replace(",", "").replace("$", "")))
            except (ValueError, TypeError):
                fmr4 = None

        if fmr3 and fmr3 > 0:
            all_rents[zipcode] = {"fmr3br": fmr3}
            if fmr4 and fmr4 > 0:
                all_rents[zipcode]["fmr4br"] = fmr4
    except (IndexError, TypeError):
        continue

wb.close()
print(f"   Parsed {total_rows:,} data rows, found {len(all_rents):,} zip codes with 3BR FMR")

# ── Step 4: Filter to our zip codes ──
print("\n🎯 Step 4: Filtering to LA/OC listings zip codes...")
rents = {}
for z in our_zips:
    if z in all_rents:
        rents[z] = all_rents[z]

matched = len(rents)
unmatched = our_zips - set(rents.keys())
print(f"   Matched: {matched}/{len(our_zips)} zip codes")
if unmatched:
    print(f"   Unmatched: {len(unmatched)} zips (no SAFMR data): {sorted(list(unmatched))[:20]}...")

# ── Step 5: Write rents.json ──
print(f"\n📦 Step 5: Writing {output_file}...")
with open(output_file, "w") as f:
    json.dump(rents, f, separators=(",", ":"))

size_kb = os.path.getsize(output_file) / 1024
print(f"   Created {output_file} ({size_kb:.1f} KB, {len(rents)} zip codes)")

# ── Summary ──
if rents:
    fmr3_vals = [v["fmr3br"] for v in rents.values() if "fmr3br" in v]
    fmr4_vals = [v["fmr4br"] for v in rents.values() if "fmr4br" in v]
    fmr3_vals.sort()
    print(f"\n📊 Summary:")
    print(f"   Total zips matched: {len(rents)}")
    print(f"   3BR FMR — Median: ${statistics.median(fmr3_vals):,.0f} | Min: ${min(fmr3_vals):,} | Max: ${max(fmr3_vals):,}")
    if fmr4_vals:
        print(f"   4BR FMR — Median: ${statistics.median(fmr4_vals):,.0f} | Min: ${min(fmr4_vals):,} | Max: ${max(fmr4_vals):,}")
    print(f"   Done! ✅\n")
else:
    print("\n   ⚠️  No matching rent data found. Check SAFMR file format.\n")
